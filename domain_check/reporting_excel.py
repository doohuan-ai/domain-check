"""Excel 报告：纵向每 URL 一行、表头与状态列/截图列配色、探针列、截图 TwoCellAnchor 随单元格隐藏。"""

from __future__ import annotations

from pathlib import Path
from typing import Mapping, Sequence

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.units import pixels_to_EMU
from domain_check.browser_check import UrlCheckResult, format_cell_status
from domain_check.config import AppConfig
from domain_check.probe_net import ProbeSummary

GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
SKIP_FILL = PatternFill("solid", fgColor="DEDEDE")

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_THIN = Side(style="thin", color="FFBFBFBF")
BORDER_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_BOTTOM = Side(style="medium", color="FF2F5597")
BORDER_HEADER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_HEADER_BOTTOM)

DATA_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
DATA_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

_IMAGE_PAD_PX = 8


def _fill_for_result(result: UrlCheckResult) -> PatternFill:
    if result.label == "success":
        return GREEN
    if result.label in ("blocked", "challenge"):
        return YELLOW
    if result.label == "skipped":
        return SKIP_FILL
    return RED


def _probe_state_label(ps: ProbeSummary) -> str:
    m = {
        "off": "关闭",
        "empty": "无 URL",
        "ok": "正常",
        "partial": "部分失败",
        "fail": "失败",
    }
    return m.get(ps.state, ps.state)


def _probe_merged_cell(ps: ProbeSummary) -> str:
    """探针摘要（不再前置“正常/部分失败/失败”文字）。"""
    d = (ps.detail or "").strip()
    if ps.state == "off":
        return "—"
    if ps.state == "empty":
        return d or "—"
    if not d:
        return _probe_state_label(ps)
    return d


def _with_extra_blank_lines(text: str) -> str:
    """
    统一把列内分隔改成「换行 + 空行」：
    - `` | `` -> 换行
    - 已有换行也扩展为双倍行距
    """
    t = (text or "").strip()
    if not t:
        return t
    t = t.replace(" | ", "\n").replace("\r\n", "\n")
    parts = [p for p in t.split("\n") if p]
    return "\n\n".join(parts)


def _precheck_detail_for_excel(text: str) -> str:
    """
    预检详情保留单行内的三段格式：
    DNS | 正常 | 17ms
    TCP:443 | 正常 | 23ms
    仅在条目之间增加空行，便于阅读。
    """
    t = (text or "").strip().replace("\r\n", "\n")
    if not t:
        return t
    if "\n" in t:
        lines = [line.strip() for line in t.split("\n") if line.strip()]
        return "\n\n".join(lines)
    # 兼容旧格式：单行用 " | " 串接多个条目时，按条目分行并空一行
    return t.replace(" | ", "\n\n")


def _probe_text_for_excel(ps: ProbeSummary) -> str:
    """探针列在 `` | `` 处分行并增加空行。"""
    raw = _probe_merged_cell(ps)
    return _with_extra_blank_lines(raw)


def _probe_profile_for_excel(ps: ProbeSummary) -> str:
    """探针链路画像列（来自 probe.urls 的 trace 等响应，英文键名 ip/loc/colo/…）。"""
    p = (ps.profile or "").strip()
    if not p:
        return "—"
    return p.replace(" | ", "\n")


def _egress_verify_for_excel(ps: ProbeSummary) -> str:
    """出口校验：外显 IP 服务 + 与路由器本批对比；并说明与页面请求同源。"""
    if ps.state == "off":
        return "—"
    ev = (ps.egress_verify or "").strip()
    note = (
        "说明：「公网IP」= 路由器本批 SNAT 目标；urllib 探针与下方页面同一台电脑、"
        "同一默认路由、顺序执行，本行所有 URL 与截图共用该出口（非逐 URL 换 IP）。"
    )
    if not ev:
        return note
    return _with_extra_blank_lines(ev) + "\n\n" + note


def _result_text_for_excel(result: UrlCheckResult) -> str:
    """结果列把 `` | `` 分隔改为换行并增加空行。"""
    return _with_extra_blank_lines(format_cell_status(result))


def _px_to_row_height_points(px: float) -> float:
    return px * 72.0 / 96.0


def _screenshot_display_size_px(cfg: AppConfig, target_h: int) -> tuple[int, int]:
    """视口比例下嵌入 Excel 的截图显示宽高（与浏览器视口一致时各图同尺寸）。"""
    vw = max(1, cfg.browser.viewport_width)
    vh = max(1, cfg.browser.viewport_height)
    tw = max(1, round(vw * target_h / vh))
    return tw, target_h


def _chars_for_screenshot_col(tw_px: int) -> float:
    """列宽（字符）≈ 容纳 tw_px + 左右留白。"""
    inner = tw_px + 2 * _IMAGE_PAD_PX + 4
    return min(92.0, max(18.0, inner / 7.0 + 2.5))


def build_workbook(
    cfg: AppConfig,
    rows: Sequence[tuple[str, list[UrlCheckResult]]],
    url_headers: list[str],
    probe_by_pub_ip: Mapping[str, ProbeSummary] | None = None,
) -> Workbook:
    """
    纵向表：每行一条 URL。
    probe_by_pub_ip: 公网 IP → 结构化探针结果（urllib 层，与浏览器列分离）；未启用时传 None。
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None

    ocfg = cfg.output
    target_h = int(ocfg.embed_screenshot_max_height)
    if target_h <= 0:
        target_h = 180

    tw_disp, th_disp = _screenshot_display_size_px(cfg, target_h)
    shot_col_wch = _chars_for_screenshot_col(tw_disp)

    headers = [
        "公网IP",
        "URL",
        "结果",
        "线路健康度",
        "预检详情(DNS/TCP/PING)",
        "探针",
        "探针链路画像",
        "出口校验",
        "截图",
    ]
    ws.append(headers)

    # 列宽整体收窄，减少横向滚动成本，优先让用户先看到更多关键列
    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 32
    ws.column_dimensions["I"].width = shot_col_wch

    for col in range(1, 10):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border = BORDER_HEADER

    last_row = 1
    probes = probe_by_pub_ip or {}

    for pub_ip, results in rows:
        ps = probes.get(pub_ip) or ProbeSummary("off", "")
        probe_text = _probe_text_for_excel(ps)
        probe_profile = _probe_profile_for_excel(ps)
        egress_text = _egress_verify_for_excel(ps)
        for url, res in zip(url_headers, results):
            ws.append(
                [
                    pub_ip,
                    url,
                    _result_text_for_excel(res),
                    res.line_health,
                    _precheck_detail_for_excel(res.precheck_detail),
                    probe_text,
                    probe_profile,
                    egress_text,
                    "",
                ]
            )
            row_num = ws.max_row
            last_row = row_num
            # 整行使用与「结果」列一致的状态底色（正常/受限/验证墙/失败/跳过）
            status_fill = _fill_for_result(res)

            for col, al in (
                (1, DATA_ALIGN_LEFT),
                (2, DATA_ALIGN_LEFT),
                (3, DATA_ALIGN_LEFT),
                (4, DATA_ALIGN_CENTER),
                (5, DATA_ALIGN_LEFT),
                (6, DATA_ALIGN_LEFT),
                (7, DATA_ALIGN_LEFT),
                (8, DATA_ALIGN_LEFT),
                (9, DATA_ALIGN_CENTER),
            ):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = status_fill
                cell.alignment = al
                cell.border = BORDER_THIN

            sp = res.screenshot_path
            if sp and Path(sp).is_file():
                p = Path(sp)
                try:
                    img = XLImage(str(p))
                    img.width = tw_disp
                    img.height = th_disp
                    h_px_total = th_disp + 2 * _IMAGE_PAD_PX
                    ws.row_dimensions[row_num].height = _px_to_row_height_points(h_px_total)
                    # twoCell：使用像素偏移定义非零锚区，避免零宽/零高导致 Excel 不显示图片
                    r0 = row_num - 1
                    img.anchor = TwoCellAnchor(
                        editAs="twoCell",
                        _from=AnchorMarker(
                            col=8,
                            colOff=pixels_to_EMU(_IMAGE_PAD_PX),
                            row=r0,
                            rowOff=pixels_to_EMU(_IMAGE_PAD_PX),
                        ),
                        to=AnchorMarker(
                            col=8,
                            colOff=pixels_to_EMU(_IMAGE_PAD_PX + tw_disp),
                            row=r0,
                            rowOff=pixels_to_EMU(_IMAGE_PAD_PX + th_disp),
                        ),
                    )
                    ws.add_image(img)
                except (OSError, ValueError, ImportError):
                    ws.row_dimensions[row_num].height = float(ocfg.data_row_height)
            else:
                ws.row_dimensions[row_num].height = float(ocfg.data_row_height)

    ws.freeze_panes = "A2"
    if last_row >= 1:
        ws.auto_filter.ref = f"A1:I{last_row}"

    return wb
