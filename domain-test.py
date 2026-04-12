"""
已废弃：请使用 Playwright 版巡检工具。

推荐入口：在项目根目录执行
  python -m domain_test -d domains.txt --output ./reports
详见 README.md。
"""

import paramiko
import subprocess
import time
import os
import tempfile
import pytesseract
import cv2
import pyautogui
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill

# ====== 写死 Tesseract 路径 ======
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# ================= 配置 =================
ROS_IP = "192.168.11.254"
ROS_PORT = 22100
ROS_USER = "shenytong"
ROS_PASS = "sronghez"

TARGET_SRC = "192.168.11.115"

CHROME_PATH = r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"

DOMAINS = [
    "https://accounts.x.ai/check-login?redirect=grok-com",
    "https://www.waterwipes.com",
    "https://app.navosagent.ai",
    "https://code.visualstudio.com",
    "https://youmind.com"
]

# Excel颜色
GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFF2CC")

# ================= SSH =================
def ssh_exec(cmd):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(ROS_IP, port=ROS_PORT, username=ROS_USER, password=ROS_PASS)
    stdin, stdout, stderr = ssh.exec_command(cmd)
    result = stdout.read().decode('gbk', errors='ignore')
    ssh.close()
    return result

def get_lo_ips():
    output = ssh_exec('/ip address print detail where interface=lo')
    ip_list = []
    ip = None
    disabled = False

    for line in output.split("\n"):
        if "address=" in line:
            ip = line.split("address=")[1].split("/")[0]
        if "disabled" in line:
            disabled = True

        if ip:
            if not disabled:
                ip_list.append(ip)
            ip = None
            disabled = False

    return ip_list

def change_nat(ip):
    print(f"切换 NAT → {ip}")
    ssh_exec(f'/ip firewall nat set [find where src-address="{TARGET_SRC}"] to-addresses={ip}')
    ssh_exec('/ip firewall connection remove [find]')
    time.sleep(8)

# ================= OCR识别 =================
def ocr_detect(img_path):
    img = cv2.imread(img_path)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # 提高识别率
    _, thresh = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY)

    text = pytesseract.image_to_string(thresh).lower()

    # 判断规则（可自行扩展）
    if any(k in text for k in ["403", "forbidden", "access denied"]):
        return "❌403/拒绝"
    if any(k in text for k in ["not available", "region", "restricted", "地区"]):
        return "⚠地区限制"
    if any(k in text for k in ["error", "failed", "timeout"]):
        return "❌访问失败"

    return "✅正常"

# ================= 浏览器截图 =================
def open_and_capture(url, img_path):
    try:
        subprocess.Popen([CHROME_PATH, "--incognito", "--new-window", url])
        time.sleep(8)

        screenshot = pyautogui.screenshot()
        screenshot.save(img_path)

    finally:
        subprocess.call("taskkill /f /im chrome.exe", shell=True)
        time.sleep(2)

# ================= Excel =================
def set_color(cell, status):
    if "✅" in status:
        cell.fill = GREEN
    elif "⚠" in status:
        cell.fill = YELLOW
    else:
        cell.fill = RED

# ================= 主流程 =================
def main():
    ip_list = get_lo_ips()

    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    file_path = os.path.join(desktop, f"OCR检测_{int(time.time())}.xlsx")

    wb = Workbook()
    ws = wb.active

    headers = ["公网IP"] + DOMAINS
    ws.append(headers)

    temp_images = []

    row = 2

    for ip in ip_list:
        print(f"\n===== 测试IP: {ip} =====")

        change_nat(ip)
        ws.cell(row=row, column=1, value=ip)

        col = 2

        for domain in DOMAINS:
            print("访问:", domain)

            img_path = os.path.join(tempfile.gettempdir(), f"{row}_{col}.png")

            open_and_capture(domain, img_path)

            status = ocr_detect(img_path)
            print("结果:", status)

            cell = ws.cell(row=row, column=col, value=status)
            set_color(cell, status)

            try:
                img = XLImage(img_path)
                img.width = 300
                img.height = 180
                ws.add_image(img, ws.cell(row=row, column=col).coordinate)
                temp_images.append(img_path)
            except:
                pass

            col += 1

        row += 1

    # 保存 Excel
    wb.save(file_path)

    # 删除临时截图
    for img in temp_images:
        try:
            os.remove(img)
        except:
            pass

    print(f"\n✅ 完成：{file_path}")

if __name__ == "__main__":
    main()