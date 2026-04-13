"""Excel 报告：纵向每 URL 一行、表头与状态列/截图列配色、嵌入截图居中留白。"""

from __future__ import annotations

from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.units import pixels_to_EMU

from domain_test.browser_check import UrlCheckResult, format_cell_status
from domain_test.config import AppConfig

# 数据行状态底色（与常见「成功/警告/失败」一致）
GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
NEUTRAL_ROW = PatternFill("solid", fgColor="FFF7F7F7")

# 表头
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_THIN = Side(style="thin", color="FFBFBFBF")
BORDER_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_BOTTOM = Side(style="medium", color="FF2F5597")
BORDER_HEADER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_HEADER_BOTTOM)

# 数据区文字：水平左对齐 + 垂直居中（URL/结果可换行）
DATA_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
DATA_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 截图列内边距（像素，用于行高/列宽与锚点偏移）
_IMAGE_PAD_PX = 8


def _fill_for_result(result: UrlCheckResult) -> PatternFill:
    if result.label == "success":
        return GREEN
    if result.label == "blocked":
        return YELLOW
    return RED


def _px_to_row_height_points(px: float) -> float:
    """Excel 行高为磅；按 96dpi 将像素近似为磅。"""
    return px * 72.0 / 96.0


def _col_width_chars_to_px(width_chars: float) -> int:
    """列宽（字符数）→ 近似像素宽。"""
    return max(1, int(width_chars * 7 + 5))


def _row_height_points_to_px(points: float) -> int:
    """行高（磅）→ 近似像素高。"""
    return max(1, int(points * 96.0 / 72.0))


def _embedded_pixel_size(image_path: Path, target_height_px: int) -> tuple[int, int]:
    """
    嵌入图固定高度 target_height_px，宽度按原图宽高比缩放。
    返回 (width_px, height_px)。
    """
    try:
        from PIL import Image as PILImage

        with PILImage.open(image_path) as im:
            w, h = im.size
        if h <= 0:
            return target_height_px, target_height_px
        tw = max(1, round(w * target_height_px / h))
        return tw, target_height_px
    except Exception:
        return target_height_px, target_height_px


def _apply_image_anchor_centered(
    img: XLImage,
    *,
    row_num: int,
    col_idx_zero: int,
    tw: int,
    th: int,
    col_width_chars: float,
    row_height_pt: float,
) -> None:
    """将图片以 OneCellAnchor 放入单元格，在可用区域内水平垂直居中（近似像素）。"""
    cell_w_px = _col_width_chars_to_px(col_width_chars)
    cell_h_px = _row_height_points_to_px(row_height_pt)
    col_off_px = max(_IMAGE_PAD_PX, (cell_w_px - tw) // 2)
    row_off_px = max(_IMAGE_PAD_PX, (cell_h_px - th) // 2)
    marker = AnchorMarker(
        col=col_idx_zero,
        colOff=pixels_to_EMU(col_off_px),
        row=row_num - 1,
        rowOff=pixels_to_EMU(row_off_px),
    )
    ext = XDRPositiveSize2D(cx=pixels_to_EMU(tw), cy=pixels_to_EMU(th))
    img.anchor = OneCellAnchor(_from=marker, ext=ext)


def build_workbook(
    cfg: AppConfig,
    rows: Sequence[tuple[str, list[UrlCheckResult]]],
    url_headers: list[str],
) -> Workbook:
    """
    纵向表：每行一条 URL（同一公网 IP 可占多行）。
    rows: (公网 IP, 与 url_headers 顺序对齐的 UrlCheckResult 列表)
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None

    ocfg = cfg.output
    target_h = int(ocfg.embed_screenshot_max_height)
    if target_h <= 0:
        target_h = 180

    headers = ["公网IP", "URL", "结果", "截图"]
    ws.append(headers)

    # 列宽初值
    for col_letter, width in (("A", 16), ("B", 52), ("C", 44), ("D", 22)):
        ws.column_dimensions[col_letter].width = width

    # 表头行（第 1 行）
    for col in range(1, 5):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border = BORDER_HEADER

    last_row = 1

    def _bump_screenshot_col_width_for_image(tw_px: int) -> None:
        inner = tw_px + 2 * _IMAGE_PAD_PX
        wch = min(92.0, max(20.0, inner / 7.0 + 4.0))
        cur = ws.column_dimensions["D"].width or 22.0
        if wch > cur:
            ws.column_dimensions["D"].width = wch

    for pub_ip, results in rows:
        for url, res in zip(url_headers, results):
            ws.append([pub_ip, url, format_cell_status(res), ""])
            row_num = ws.max_row
            last_row = row_num
            status_fill = _fill_for_result(res)

            for col, al, fill in (
                (1, DATA_ALIGN_LEFT, NEUTRAL_ROW),
                (2, DATA_ALIGN_LEFT, NEUTRAL_ROW),
                (3, DATA_ALIGN_LEFT, status_fill),
                (4, DATA_ALIGN_CENTER, status_fill),
            ):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = fill
                cell.alignment = al
                cell.border = BORDER_THIN

            sp = res.screenshot_path
            if sp and Path(sp).is_file():
                p = Path(sp)
                try:
                    img = XLImage(str(p))
                    tw, th = _embedded_pixel_size(p, target_h)
                    img.width = tw
                    img.height = th
                    _bump_screenshot_col_width_for_image(tw)

                    h_px_total = th + 2 * _IMAGE_PAD_PX
                    ws.row_dimensions[row_num].height = _px_to_row_height_points(h_px_total)

                    d_w = float(ws.column_dimensions["D"].width or 22.0)
                    r_h = float(ws.row_dimensions[row_num].height)

                    _apply_image_anchor_centered(
                        img,
                        row_num=row_num,
                        col_idx_zero=3,
                        tw=tw,
                        th=th,
                        col_width_chars=d_w,
                        row_height_pt=r_h,
                    )
                    ws.add_image(img)
                except OSError:
                    ws.row_dimensions[row_num].height = float(ocfg.data_row_height)
            else:
                ws.row_dimensions[row_num].height = float(ocfg.data_row_height)

    ws.freeze_panes = "A2"
    if last_row >= 1:
        ws.auto_filter.ref = f"A1:D{last_row}"

    return wb
